import openpyxl

# Path to the MCA RFID Sheet
xlsx_path = r'c:\Users\ogita\OneDrive\Desktop\rfid_website_new\MCA RFID Sheet 2025.xlsx'

# Load workbook and sheet
wb = openpyxl.load_workbook(xlsx_path)
sheet = wb.active

print("Excel file structure:")
print("First 5 rows:")
for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=5, values_only=True), 1):
    print(f"Row {i}: {row}")

print("\nColumn headers (first row):")
headers = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
for i, header in enumerate(headers):
    print(f"Column {i}: {header}")